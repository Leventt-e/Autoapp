
import toga
from toga.style import Pack
import openpyxl
import pandas as pd

class Autok:
    def __init__(self, tipus, evjarat, kilometer, ar):
        self.tipus = tipus
        self.evjarat = evjarat
        self.kilometer = kilometer
        self.ar = ar
def import_autok(excel_file):
    # Excel-fájl beolvasása pandas segítségével
    df = pd.read_excel(excel_file)

    # Az importált táblázat sorainak bejárása és Autok objektumok létrehozása
    auto_keszlet_list = []
    for index, row in df.iterrows():
        autok = Autok(tipus=row['tipus'], evjarat=row['evjarat'], kilometer=row['kilometer'], ar=row['ar'])
        auto_keszlet_list.append(autok)
    return auto_keszlet_list

def max(widget):
    excel_file_path = 'C:/temp/autok.xlsx'
    auto_keszlet = import_autok(excel_file_path)
    minimum_kor = 3000 # ettől az ávszámtól biztos, hogy kisebb lesz a legöregebb autó életkora

    # Kiírjuk az importált adatokat
    for autok in auto_keszlet:
        print(f"Tipus: {autok.tipus}, Évjárat: {autok.evjarat}, Kilométer: {autok.evjarat}, Ár: {autok.ar}")
        if int(autok.evjarat) < minimum_kor:
            minimum_kor = int(autok.evjarat)

        #maxi = (f"Tipus: {minimum_kor.tipus}, Évjárat: {minimum_kor.evjarat}, Kilométer: {minimum_kor.evjarat}, Ár: {minimum_kor.ar}")

        #if autok.evjarat == 2000:
            #maxi = autok.tipus
            #maxi = ("típusa:",autok.tipus,"évjárata:",autok.evjarat,"futott kilométer:",autok.kilometer," ár:",autok.ar)
            #maxi = (f"Tipus: {minimum_kor.tipus}, Évjárat: {minimum_kor.evjarat}, Kilométer: {minimum_kor.evjarat}, Ár: {minimum_kor.ar}")

    print("Legidősebb autót: ", minimum_kor, "-ban gyártották.")


# Itt kezdődok az applikáció
class autok(toga.App):
    def startup(self):
        # Az a bemeneti mező ahová kiírásra kerül a legidősebb autó életkora
        self.old_auto = toga.TextInput(readonly=True)

        # Adatok előkészítése
        self.data = []

        # Táblázat widget létrehozása
        self.table = toga.Table(
            data=self.data,
            headings=['Típus', 'Évjárat', 'Kilométer', 'Ár'],
            style=Pack(flex=1, height=250)
        )

        # Az input mezők definiálása
        tipus_label = toga.Label("Típus:")
        tipus_label.style.update(padding_left=27, padding_top=5, padding_bottom=3)
        self.tipus_input = toga.TextInput()
        self.tipus_input.style.update(width=150, padding_left=30, padding_top=2, padding_right=30, padding_bottom=3)

        evjarat_label = toga.Label("Évjárat:")
        evjarat_label.style.update(flex=1, padding_left=27, padding_top=5, padding_bottom=3)
        self.evjarat_input = toga.TextInput()
        self.evjarat_input.style.update(flex=1, width=150, padding_left=30, padding_top=2, padding_right=30, padding_bottom=3)

        kilometer_label = toga.Label("Kilométer:")
        kilometer_label.style.update(flex=1, padding_left=27, padding_top=5, padding_bottom=3)
        self.kilometer_input = toga.TextInput()
        self.kilometer_input.style.update(flex=1, width=150, padding_left=30, padding_top=2, padding_right=30, padding_bottom=3)

        ar_label = toga.Label("Ár:")
        ar_label.style.update(flex=1, padding_left=27, padding_top=5, padding_bottom=3)
        self.ar_input = toga.TextInput()
        self.ar_input.style.update(flex=1,  width=150,padding_left=30, padding_top=2, padding_right=30, padding_bottom=3)

        old_auto_label = toga.Label("A legidősebb autó életkora:")
        old_auto_label.style.update(flex=1, padding_left=27, padding_top=5, padding_bottom=3)
        self.old_auto_input = toga.TextInput()
        self.old_auto.style.update(flex=1,  width=150,padding_left=30, padding_top=2, padding_right=30, padding_bottom=3)

        # Ablak elrendezés összeállítása
        main_box = toga.Box(
            children=[
                self.table,
                tipus_label,
                self.tipus_input,
                evjarat_label,
                self.evjarat_input,
                kilometer_label,
                self.kilometer_input,
                ar_label,
                self.ar_input,
                old_auto_label,
                self.old_auto,
                ],
            style=Pack(direction='column', padding=10)
        )

        # A menüpontok létrehozása
        cmd1 = toga.Command(
            self.import_tablazatbol,
            text="Rögzített adatok megtekintése",
            tooltip="Az excel táblázatban lévő valamennyi adat megtekintése",
            icon=toga.Icon.TOGA_ICON,
        )
        cmd2 = toga.Command(
            self.export_tablazatba,
            text="Autók megjelenítése és rögzítése",
            tooltip="Autók megjelenítése és az adatok exportálása az excel táblázatba.",
            icon=toga.Icon.TOGA_ICON,
        )
        cmd3 = toga.Command(
            max,
            text="Legidősebb autó életkora",
            tooltip="Autók megjelenítése és az adatok exportálása az excel táblázatba.",
            icon=toga.Icon.TOGA_ICON,
        )

        # Ablak létrehozása, kiírása
        self.main_window = toga.MainWindow(title=self.formal_name, size=(650, 400))
        self.main_window.content = main_box
        self.main_window.toolbar.add(cmd1, cmd2, cmd3)

        # Beolvassa és megjeleníti az adatbázisban lévő adatokat
        self.import_tablazatbol(widget=main_box)
        self.main_window.show()

    # Az adatok importálása a táblázatból
    def import_tablazatbol(self, widget):
        file_path = "C:/temp/autok.xlsx"
        if file_path:
            # Adatok beolvasása Excel fájlból
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            imported_data = [list(row) for row in sheet.iter_rows(values_only=True)]

            # Ellenőrzés: legalább egy sor van-e
            if not imported_data:
                self.main_window.info_dialog('Hiba', 'Nem található adat az Excel fájlban.')
                return

            # Ellenőrzés: minden sor ugyanannyi oszlopot tartalmaz-e, mint az eredeti táblázat
            if len(imported_data[0]) != len(self.table.headings):
                self.main_window.info_dialog('Hiba',
                                             'Az Excel fájl oszlopainak száma nem egyezik a táblázattal.')
                return

            # Importált adatok hozzáadása a táblázathoz
            self.data.extend(imported_data)
            self.table.data = self.data

# Az adatok exportálása a táblázatba
    def export_tablazatba(self, widget):
        # Felhasználói adatok
        tipus = self.tipus_input.value
        evjarat = self.evjarat_input.value
        kilometer = self.kilometer_input.value
        ar = self.ar_input.value

        # Üres mezők ellenőrzése
        if not tipus or not evjarat or not kilometer or not ar:
            self.main_window.info_dialog('Figyelem!', 'Üres mező nem lehet!')
            return

        # Felhasználói adatok hozzáadása a táblázathoz
        self.data.append([tipus, evjarat, int(kilometer), ar])
        self.table.data = self.data

        # Input mezők ürítése
        self.tipus_input.value = ''
        self.evjarat_input.value = ''
        self.kilometer_input.value = ''
        self.ar_input.value = ''
        #self.telephely_input.value = ''

        # XLSX fájl létrehozása és adatok írása
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # Oszlopfejlékek írása
        headings = self.data[0]
        for col_num, heading in enumerate(headings, 1):
            sheet.cell(row=1, column=col_num, value=heading)

        # Adatok írása
        for row_num, row_data in enumerate(self.data[1:], 2):
            for col_num, cell_value in enumerate(row_data, 1):
                sheet.cell(row=row_num, column=col_num, value=cell_value)

        # Fájl mentése
        file_path = "C:/temp/autok.xlsx"
        if file_path:
            workbook.save(file_path)

        # Üzenetablak megjelenítése a sikeres mentésről
        self.main_window.info_dialog('O.K.', 'Az exportálás sikeres!')

def main():
    return autok()
